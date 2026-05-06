"""Ingest the Artsy Match Workbook (xlsx).

Expected `Summary` sheet columns:
    Inventory ID, Title, Classification, Materials,
    Primary Source Kind, Primary Source PDF, Primary Page,
    Primary Candidate, Primary Confidence, Primary Note, Thumbnail

Each row maps a KG-# to a primary image source PDF + page + confidence.
We emit observations for `title`, `classification`, `materials`, plus a
"primary_image_source" field that records "PDF p.N (confidence)".
"""

from __future__ import annotations

from openpyxl import load_workbook

from ..ids import parse_kg_id
from ..normalize import (
    clean, normalize_classification, normalize_title,
)
from ..schema import Observation
from ._base import IngestResult, Ingester


CONF_MAP = {"high": "high", "medium": "medium", "low": "low"}


class MatchWorkbookIngester(Ingester):
    type = "match_workbook"

    def run(self) -> IngestResult:
        observations: list[Observation] = []
        rows_processed = 0
        wb = load_workbook(filename=str(self.file_path), read_only=True, data_only=True)
        sheet = wb["Summary"] if "Summary" in wb.sheetnames else wb.active

        header_row = None
        col_idx: dict[str, int] = {}
        for r_i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if header_row is None:
                if not row or not any(row):
                    continue
                if any(isinstance(c, str) and "Inventory ID" in c for c in row if c):
                    header_row = r_i
                    for i, c in enumerate(row):
                        if c is None:
                            continue
                        col_idx[str(c).strip()] = i
                continue
            rows_processed += 1
            kg = parse_kg_id(str(row[col_idx.get("Inventory ID", 0)] or ""))
            if not kg:
                continue
            title = normalize_title(row[col_idx.get("Title", -1)] if "Title" in col_idx else None)
            cls = normalize_classification(
                row[col_idx.get("Classification", -1)] if "Classification" in col_idx else None
            )
            materials = clean(row[col_idx.get("Materials", -1)] if "Materials" in col_idx else None)
            pdf = clean(row[col_idx.get("Primary Source PDF", -1)] if "Primary Source PDF" in col_idx else None)
            page = clean(str(row[col_idx.get("Primary Page", -1)]) if "Primary Page" in col_idx and row[col_idx["Primary Page"]] is not None else None)
            conf = (clean(row[col_idx.get("Primary Confidence", -1)] if "Primary Confidence" in col_idx else None) or "medium").lower()
            confidence = CONF_MAP.get(conf, "medium")

            ref = f"row={r_i}"
            for field, value in (
                ("title", title),
                ("classification", cls),
                ("materials", materials),
            ):
                if value:
                    observations.append(Observation(
                        work_id=kg, field=field, value=str(value),
                        source_row_ref=ref, confidence=confidence,
                    ))
            if pdf:
                summary = f"{pdf} p.{page}" if page else pdf
                observations.append(Observation(
                    work_id=kg, field="primary_image_source", value=summary,
                    source_row_ref=ref, confidence=confidence,
                ))

        wb.close()
        return IngestResult(
            source=self._make_source(row_count=rows_processed),
            observations=observations,
        )
